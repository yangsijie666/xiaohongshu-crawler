"""
数据存储模块

职责：
  - 将采集结果持久化到本地文件
  - 支持两种格式：JSON（原始完整数据）和 Excel（多 Sheet 表格）
  - 按关键词和时间戳组织文件命名，避免覆盖

目录结构：
    data/
    ├── raw/
    │   ├── {keyword}_{timestamp}.json         # 搜索结果原始数据
    │   └── notes_{keyword}_{timestamp}.json   # 笔记详情原始数据
    └── processed/
        └── {keyword}_{timestamp}.xlsx         # Excel 汇总（3 个 Sheet）

用法：
    storage = Storage(config["storage"])
    storage.save_all("Python教程", search_results, note_details)
"""

from __future__ import annotations

import json
import logging
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Excel 各 Sheet 的列头定义
_SEARCH_FIELDS = [
    "note_id",
    "title",
    "author",
    "author_id",
    "likes",
    "note_type",
    "note_url",
    "publish_time",
]

_NOTE_FIELDS = [
    "note_id",
    "title",
    "content",
    "author",
    "author_id",
    "publish_time",
    "likes",
    "collects",
    "comments_count",
    "shares",
    "tags",
    "note_type",
    "note_url",
]

_COMMENT_FIELDS = [
    "comment_id",
    "note_id",
    "user_name",
    "user_id",
    "content",
    "likes",
    "time",
    "ip_location",
]


class Storage:
    """本地数据存储管理器。

    根据配置决定是否写入 JSON / Excel，负责目录创建和文件命名。
    """

    def __init__(self, config: dict) -> None:
        """初始化存储配置。

        Args:
            config: settings.yaml 中 storage 节点的字典，包含：
                - output_dir (str): 输出根目录，默认 "data"
                - save_raw_json (bool): 是否保存原始 JSON
                - save_xlsx (bool): 是否保存 Excel
        """
        self._root = Path(config.get("output_dir", "data"))
        self._save_json: bool = config.get("save_raw_json", True)
        self._save_xlsx: bool = config.get("save_xlsx", True)
        self._ensure_dirs()

    def _ensure_dirs(self) -> None:
        """确保输出目录存在。"""
        (self._root / "raw").mkdir(parents=True, exist_ok=True)
        (self._root / "processed").mkdir(parents=True, exist_ok=True)

    def save_all(
        self,
        keyword: str,
        search_results: list[dict],
        note_details: list[dict],
    ) -> None:
        """统一保存所有采集数据（JSON + Excel）。

        Args:
            keyword: 搜索关键词（用于文件命名）
            search_results: parse_search_card() 返回的字典列表
            note_details: fetch_note_details() 返回的笔记详情列表，
                          每条包含详情字段 + comments 子列表
        """
        safe_keyword = _sanitize_filename(keyword)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        # JSON 写入
        if self._save_json:
            if search_results:
                self._write_json(safe_keyword, timestamp, keyword, search_results)
            if note_details:
                self._write_notes_json(safe_keyword, timestamp, keyword, note_details)

        # Excel 写入
        if self._save_xlsx:
            self._write_xlsx(safe_keyword, timestamp, search_results, note_details)

    # ---- JSON 写入方法 ----

    def _write_json(
        self,
        safe_keyword: str,
        timestamp: str,
        keyword: str,
        results: list[dict],
    ) -> None:
        """写入搜索结果 JSON 文件。"""
        json_path = self._root / "raw" / f"{safe_keyword}_{timestamp}.json"
        payload = {
            "keyword": keyword,
            "crawled_at": datetime.now().isoformat(timespec="seconds"),
            "count": len(results),
            "results": results,
        }
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
        logger.info("JSON 已写入：%s（%d 条）", json_path, len(results))

    def _write_notes_json(
        self,
        safe_keyword: str,
        timestamp: str,
        keyword: str,
        note_details: list[dict],
    ) -> None:
        """写入笔记详情 JSON 文件（含评论）。"""
        json_path = self._root / "raw" / f"notes_{safe_keyword}_{timestamp}.json"
        payload = {
            "keyword": keyword,
            "crawled_at": datetime.now().isoformat(timespec="seconds"),
            "count": len(note_details),
            "notes": note_details,
        }
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
        logger.info("笔记详情 JSON 已写入：%s（%d 条）", json_path, len(note_details))

    # ---- Excel 写入方法 ----

    def _write_xlsx(
        self,
        safe_keyword: str,
        timestamp: str,
        search_results: list[dict],
        note_details: list[dict],
    ) -> None:
        """生成包含 3 个 Sheet 的 Excel 文件。

        Sheet 结构：
          - 搜索结果：搜索阶段获取的笔记摘要
          - 笔记详情：笔记正文、互动数据等
          - 评论：所有笔记的评论汇总
        """
        wb = Workbook()

        # Sheet 1: 搜索结果
        ws_search = wb.active
        ws_search.title = "搜索结果"
        self._fill_sheet(ws_search, _SEARCH_FIELDS, search_results)

        # Sheet 2: 笔记详情（tags 列表转字符串，移除嵌套字段）
        ws_notes = wb.create_sheet("笔记详情")
        note_rows = []
        for note in note_details:
            row = dict(note)
            row["tags"] = ";".join(row.get("tags", []))
            row.pop("comments", None)
            row.pop("images", None)
            row.pop("video_url", None)
            note_rows.append(row)
        self._fill_sheet(ws_notes, _NOTE_FIELDS, note_rows)

        # Sheet 3: 评论汇总
        ws_comments = wb.create_sheet("评论")
        all_comments: list[dict] = []
        for note in note_details:
            all_comments.extend(note.get("comments", []))
        self._fill_sheet(ws_comments, _COMMENT_FIELDS, all_comments)

        # 保存文件
        xlsx_path = self._root / "processed" / f"{safe_keyword}_{timestamp}.xlsx"
        wb.save(xlsx_path)
        logger.info(
            "Excel 已写入：%s（搜索 %d 条 / 笔记 %d 条 / 评论 %d 条）",
            xlsx_path,
            len(search_results),
            len(note_details),
            len(all_comments),
        )

    def _fill_sheet(
        self,
        ws,
        fieldnames: list[str],
        rows: list[dict],
    ) -> None:
        """填充单个 Sheet：写入表头 + 数据行 + 格式化。

        格式化包括：冻结首行、自动筛选、自适应列宽。
        """
        # 写入表头
        ws.append(fieldnames)

        # 写入数据行
        for row in rows:
            ws.append([row.get(field) for field in fieldnames])

        # 冻结首行（滚动时表头始终可见）
        ws.freeze_panes = "A2"

        # 自动筛选（覆盖所有数据列）
        if rows:
            last_col = get_column_letter(len(fieldnames))
            last_row = len(rows) + 1  # +1 表头行
            ws.auto_filter.ref = f"A1:{last_col}{last_row}"

        # 自适应列宽（基于表头和内容的最大长度）
        for col_idx, field in enumerate(fieldnames, start=1):
            # 计算该列最大字符宽度（表头 + 前 100 行数据取样）
            max_len = len(str(field))
            for row in rows[:100]:
                val = row.get(field)
                if val is not None:
                    # 中文字符按 2 倍宽度计算
                    cell_len = sum(2 if ord(c) > 127 else 1 for c in str(val))
                    max_len = max(max_len, cell_len)
            # 限制最大列宽为 60，最小为 10
            col_width = min(max(max_len + 2, 10), 60)
            ws.column_dimensions[get_column_letter(col_idx)].width = col_width


def _sanitize_filename(name: str) -> str:
    """将字符串转化为安全的文件名（去除 / \\ : * ? " < > | 等特殊字符）。

    Args:
        name: 原始字符串

    Returns:
        安全的文件名字符串（保留中文、字母、数字、下划线、连字符）
    """
    import re
    # 替换不安全字符为下划线
    safe = re.sub(r'[\\/:*?"<>|\s]', "_", name)
    # 合并连续下划线
    safe = re.sub(r"_+", "_", safe)
    return safe.strip("_") or "unnamed"
