"""
Storage 模块单元测试

测试策略：
  - 使用 pytest tmp_path fixture 隔离文件 I/O，不污染真实数据目录
  - 覆盖 Storage 类的完整行为：目录创建、JSON 写入、Excel 写入
  - 覆盖 _sanitize_filename 纯函数的各种输入格式
"""

from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from src.storage import Storage, _sanitize_filename

# ---- 测试数据 ----

SAMPLE_SEARCH_RESULTS = [
    {
        "note_id": "abc123",
        "title": "Python 入门教程",
        "author": "小红书用户",
        "author_id": "user001",
        "likes": 1200,
        "note_type": "image",
        "note_url": "https://www.xiaohongshu.com/explore/abc123",
        "publish_time": "2025-01-15",
    },
    {
        "note_id": "def456",
        "title": "Python 进阶",
        "author": "另一位用户",
        "author_id": "user002",
        "likes": 500,
        "note_type": "video",
        "note_url": "https://www.xiaohongshu.com/explore/def456",
        "publish_time": "2025-01-16",
    },
]

SAMPLE_NOTE_DETAILS = [
    {
        "note_id": "abc123",
        "title": "Python 入门教程",
        "content": "这是一篇 Python 入门教程...",
        "author": "小红书用户",
        "author_id": "user001",
        "publish_time": "2025-01-15",
        "likes": 1200,
        "collects": 300,
        "comments_count": 50,
        "shares": 20,
        "tags": ["Python", "编程", "教程"],
        "note_type": "image",
        "note_url": "https://www.xiaohongshu.com/explore/abc123",
        "images": ["https://example.com/img1.jpg"],
        "video_url": "",
        "comments": [
            {
                "comment_id": "cmt001",
                "note_id": "abc123",
                "user_name": "评论用户",
                "user_id": "usr001",
                "content": "太有用了！",
                "likes": 10,
                "time": "01-15",
                "ip_location": "广东",
            }
        ],
    }
]


class TestSanitizeFilename:
    """测试文件名安全化纯函数。"""

    def test_normal_alphanumeric_unchanged(self):
        """普通字母数字字符串应保持不变。"""
        assert _sanitize_filename("Python123") == "Python123"

    def test_chinese_keyword_preserved(self):
        """中文字符应被保留。"""
        result = _sanitize_filename("小红书教程")
        assert "小红书教程" in result

    def test_replaces_slash(self):
        """斜杠应被替换。"""
        result = _sanitize_filename("a/b")
        assert "/" not in result

    def test_replaces_colon(self):
        """冒号应被替换。"""
        result = _sanitize_filename("a:b")
        assert ":" not in result

    def test_replaces_spaces(self):
        """空格应被替换。"""
        result = _sanitize_filename("hello world")
        assert " " not in result

    def test_merges_consecutive_underscores(self):
        """连续下划线应合并为单个。"""
        result = _sanitize_filename("a//b")
        assert "__" not in result

    def test_strips_leading_and_trailing_underscores(self):
        """首尾下划线应被去除。"""
        result = _sanitize_filename("/test/")
        assert not result.startswith("_")
        assert not result.endswith("_")

    def test_all_special_chars_returns_unnamed(self):
        """全部为不安全字符时应返回 'unnamed'。"""
        assert _sanitize_filename("///") == "unnamed"
        assert _sanitize_filename("?*<>") == "unnamed"

    def test_replaces_all_unsafe_chars(self):
        """所有不安全字符应被替换。"""
        for ch in r'\/:*?"<>|':
            result = _sanitize_filename(ch + "text" + ch)
            assert ch not in result


class TestStorageInit:
    """测试 Storage 初始化行为。"""

    def test_creates_raw_directory(self, tmp_path):
        """初始化时应创建 raw/ 子目录。"""
        Storage({"output_dir": str(tmp_path / "data")})
        assert (tmp_path / "data" / "raw").is_dir()

    def test_creates_processed_directory(self, tmp_path):
        """初始化时应创建 processed/ 子目录。"""
        Storage({"output_dir": str(tmp_path / "data")})
        assert (tmp_path / "data" / "processed").is_dir()

    def test_default_output_dir_is_data(self, tmp_path, monkeypatch):
        """未指定 output_dir 时默认使用 'data'。"""
        monkeypatch.chdir(tmp_path)
        Storage({})
        assert (tmp_path / "data" / "raw").is_dir()

    def test_save_json_defaults_to_true(self, tmp_path):
        """save_raw_json 默认为 True。"""
        s = Storage({"output_dir": str(tmp_path)})
        assert s._save_json is True

    def test_save_xlsx_defaults_to_true(self, tmp_path):
        """save_xlsx 默认为 True。"""
        s = Storage({"output_dir": str(tmp_path)})
        assert s._save_xlsx is True

    def test_save_json_can_be_disabled(self, tmp_path):
        """save_raw_json=False 可关闭 JSON 写入。"""
        s = Storage({"output_dir": str(tmp_path), "save_raw_json": False})
        assert s._save_json is False

    def test_save_xlsx_can_be_disabled(self, tmp_path):
        """save_xlsx=False 可关闭 Excel 写入。"""
        s = Storage({"output_dir": str(tmp_path), "save_xlsx": False})
        assert s._save_xlsx is False


class TestWriteJson:
    """测试 JSON 写入功能。"""

    def test_creates_search_json_file(self, tmp_path):
        """有搜索结果时应在 raw/ 目录创建 JSON 文件。"""
        storage = Storage({"output_dir": str(tmp_path), "save_xlsx": False})
        storage.save_all("Python", SAMPLE_SEARCH_RESULTS, [])
        raw_files = list((tmp_path / "raw").glob("*.json"))
        assert len(raw_files) == 1

    def test_search_json_has_correct_structure(self, tmp_path):
        """搜索 JSON 文件应包含 keyword / crawled_at / count / results。"""
        storage = Storage({"output_dir": str(tmp_path), "save_xlsx": False})
        storage.save_all("Python", SAMPLE_SEARCH_RESULTS, [])
        json_file = list((tmp_path / "raw").glob("Python_*.json"))[0]
        with open(json_file, encoding="utf-8") as f:
            data = json.load(f)
        assert data["keyword"] == "Python"
        assert "crawled_at" in data
        assert data["count"] == len(SAMPLE_SEARCH_RESULTS)
        assert len(data["results"]) == len(SAMPLE_SEARCH_RESULTS)

    def test_notes_json_created_when_details_exist(self, tmp_path):
        """有笔记详情时应创建 notes_{keyword}_*.json 文件。"""
        storage = Storage({"output_dir": str(tmp_path), "save_xlsx": False})
        storage.save_all("Python", SAMPLE_SEARCH_RESULTS, SAMPLE_NOTE_DETAILS)
        notes_files = list((tmp_path / "raw").glob("notes_Python_*.json"))
        assert len(notes_files) == 1

    def test_notes_json_contains_notes_key(self, tmp_path):
        """笔记详情 JSON 应包含 notes 键及 count。"""
        storage = Storage({"output_dir": str(tmp_path), "save_xlsx": False})
        storage.save_all("Python", SAMPLE_SEARCH_RESULTS, SAMPLE_NOTE_DETAILS)
        notes_file = list((tmp_path / "raw").glob("notes_Python_*.json"))[0]
        with open(notes_file, encoding="utf-8") as f:
            data = json.load(f)
        assert "notes" in data
        assert data["count"] == len(SAMPLE_NOTE_DETAILS)

    def test_no_json_when_search_results_empty(self, tmp_path):
        """搜索结果为空时不创建搜索 JSON 文件。"""
        storage = Storage({"output_dir": str(tmp_path), "save_xlsx": False})
        storage.save_all("Python", [], [])
        raw_files = list((tmp_path / "raw").glob("*.json"))
        assert len(raw_files) == 0

    def test_no_json_when_disabled(self, tmp_path):
        """save_raw_json=False 时不创建任何 JSON 文件。"""
        storage = Storage(
            {"output_dir": str(tmp_path), "save_raw_json": False, "save_xlsx": False}
        )
        storage.save_all("Python", SAMPLE_SEARCH_RESULTS, SAMPLE_NOTE_DETAILS)
        raw_files = list((tmp_path / "raw").glob("*.json"))
        assert len(raw_files) == 0

    def test_json_preserves_unicode(self, tmp_path):
        """中文内容应正确写入 UTF-8 JSON。"""
        storage = Storage({"output_dir": str(tmp_path), "save_xlsx": False})
        storage.save_all("小红书", SAMPLE_SEARCH_RESULTS, [])
        json_file = list((tmp_path / "raw").glob("*.json"))[0]
        content = json_file.read_text(encoding="utf-8")
        assert "Python 入门教程" in content


class TestWriteXlsx:
    """测试 Excel 写入功能。"""

    def test_creates_xlsx_file(self, tmp_path):
        """有数据时应在 processed/ 目录创建 xlsx 文件。"""
        storage = Storage({"output_dir": str(tmp_path), "save_raw_json": False})
        storage.save_all("Python", SAMPLE_SEARCH_RESULTS, SAMPLE_NOTE_DETAILS)
        xlsx_files = list((tmp_path / "processed").glob("*.xlsx"))
        assert len(xlsx_files) == 1

    def test_xlsx_has_three_sheets(self, tmp_path):
        """Excel 文件应包含 搜索结果 / 笔记详情 / 评论 三个 Sheet。"""
        storage = Storage({"output_dir": str(tmp_path), "save_raw_json": False})
        storage.save_all("Python", SAMPLE_SEARCH_RESULTS, SAMPLE_NOTE_DETAILS)
        xlsx_file = list((tmp_path / "processed").glob("*.xlsx"))[0]
        wb = load_workbook(xlsx_file)
        assert "搜索结果" in wb.sheetnames
        assert "笔记详情" in wb.sheetnames
        assert "评论" in wb.sheetnames

    def test_search_sheet_has_header_row(self, tmp_path):
        """搜索结果 Sheet 应有正确的表头行。"""
        storage = Storage({"output_dir": str(tmp_path), "save_raw_json": False})
        storage.save_all("Python", SAMPLE_SEARCH_RESULTS, [])
        xlsx_file = list((tmp_path / "processed").glob("*.xlsx"))[0]
        ws = load_workbook(xlsx_file)["搜索结果"]
        header = [cell.value for cell in ws[1]]
        assert "note_id" in header
        assert "title" in header
        assert "author" in header

    def test_search_sheet_row_count(self, tmp_path):
        """搜索结果 Sheet 的行数应为表头+数据行。"""
        storage = Storage({"output_dir": str(tmp_path), "save_raw_json": False})
        storage.save_all("Python", SAMPLE_SEARCH_RESULTS, [])
        xlsx_file = list((tmp_path / "processed").glob("*.xlsx"))[0]
        ws = load_workbook(xlsx_file)["搜索结果"]
        assert ws.max_row == 1 + len(SAMPLE_SEARCH_RESULTS)

    def test_comments_aggregated_in_third_sheet(self, tmp_path):
        """所有笔记的评论应汇总到评论 Sheet（1 表头 + 评论条数）。"""
        storage = Storage({"output_dir": str(tmp_path), "save_raw_json": False})
        storage.save_all("Python", SAMPLE_SEARCH_RESULTS, SAMPLE_NOTE_DETAILS)
        xlsx_file = list((tmp_path / "processed").glob("*.xlsx"))[0]
        ws = load_workbook(xlsx_file)["评论"]
        total_comments = sum(len(n.get("comments", [])) for n in SAMPLE_NOTE_DETAILS)
        assert ws.max_row == 1 + total_comments

    def test_sheet_has_frozen_pane(self, tmp_path):
        """表头行应设置冻结窗格（A2）。"""
        storage = Storage({"output_dir": str(tmp_path), "save_raw_json": False})
        storage.save_all("Python", SAMPLE_SEARCH_RESULTS, [])
        xlsx_file = list((tmp_path / "processed").glob("*.xlsx"))[0]
        ws = load_workbook(xlsx_file)["搜索结果"]
        assert ws.freeze_panes == "A2"

    def test_note_tags_serialized_as_semicolon_string(self, tmp_path):
        """笔记详情 Sheet 中 tags 列表应序列化为分号分隔字符串。"""
        storage = Storage({"output_dir": str(tmp_path), "save_raw_json": False})
        storage.save_all("Python", [], SAMPLE_NOTE_DETAILS)
        xlsx_file = list((tmp_path / "processed").glob("*.xlsx"))[0]
        ws = load_workbook(xlsx_file)["笔记详情"]
        header = [cell.value for cell in ws[1]]
        tags_col = header.index("tags") + 1
        tags_cell = ws.cell(row=2, column=tags_col)
        assert tags_cell.value == "Python;编程;教程"

    def test_no_xlsx_when_disabled(self, tmp_path):
        """save_xlsx=False 时不创建 xlsx 文件。"""
        storage = Storage(
            {"output_dir": str(tmp_path), "save_raw_json": False, "save_xlsx": False}
        )
        storage.save_all("Python", SAMPLE_SEARCH_RESULTS, SAMPLE_NOTE_DETAILS)
        xlsx_files = list((tmp_path / "processed").glob("*.xlsx"))
        assert len(xlsx_files) == 0

    def test_xlsx_created_with_empty_data(self, tmp_path):
        """即使数据为空，save_xlsx=True 也应创建 xlsx 文件。"""
        storage = Storage({"output_dir": str(tmp_path), "save_raw_json": False})
        storage.save_all("Python", [], [])
        xlsx_files = list((tmp_path / "processed").glob("*.xlsx"))
        assert len(xlsx_files) == 1

    def test_xlsx_column_width_is_set(self, tmp_path):
        """列宽应被自动设置（不为默认 None）。"""
        storage = Storage({"output_dir": str(tmp_path), "save_raw_json": False})
        storage.save_all("Python", SAMPLE_SEARCH_RESULTS, [])
        xlsx_file = list((tmp_path / "processed").glob("*.xlsx"))[0]
        ws = load_workbook(xlsx_file)["搜索结果"]
        # 至少第一列（note_id）的列宽应被设置
        col_width = ws.column_dimensions["A"].width
        assert col_width is not None
        assert col_width > 0
